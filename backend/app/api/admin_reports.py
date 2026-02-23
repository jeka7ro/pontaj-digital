from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_
from datetime import datetime, date
from typing import Optional
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models import Timesheet, User, ConstructionSite, TimesheetSegment, TimesheetLine
from app.api.admin_auth import get_current_admin

router = APIRouter()

@router.get("/timesheets/excel")
async def export_timesheets_excel(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    employee_id: Optional[str] = Query(None),
    site_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin)
):
    """
    Generate Excel file with timesheet data
    """
    # Build query
    query = db.query(Timesheet).join(User).join(ConstructionSite)
    
    # Apply filters
    filters = []
    if date_from:
        filters.append(Timesheet.date >= datetime.strptime(date_from, "%Y-%m-%d").date())
    if date_to:
        filters.append(Timesheet.date <= datetime.strptime(date_to, "%Y-%m-%d").date())
    if employee_id:
        filters.append(Timesheet.user_id == employee_id)
    if site_id:
        filters.append(Timesheet.site_id == site_id)
    
    if filters:
        query = query.filter(and_(*filters))
    
    # Order by date, then employee
    query = query.order_by(Timesheet.date.desc(), User.full_name)
    
    timesheets = query.all()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pontaje"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_side = Side(style='thin', color="000000")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Headers
    headers = [
        "Data", "Angajat", "Cod Angajat", "Rol", "Șantier", 
        "Intrare", "Ieșire", "Pauză (min)", "Ore Total", "Activități"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    row_num = 2
    total_hours = 0.0
    
    for timesheet in timesheets:
        # Calculate total hours
        hours_worked = 0.0
        if timesheet.check_in and timesheet.check_out:
            delta = timesheet.check_out - timesheet.check_in
            hours_worked = delta.total_seconds() / 3600
            if timesheet.break_duration:
                hours_worked -= timesheet.break_duration / 60
        
        total_hours += hours_worked
        
        # Get activities
        activities = []
        segments = db.query(TimesheetSegment).filter(
            TimesheetSegment.timesheet_id == timesheet.id
        ).all()
        
        for segment in segments:
            lines = db.query(TimesheetLine).filter(
                TimesheetLine.segment_id == segment.id
            ).all()
            for line in lines:
                if line.quantity:
                    activities.append(f"{line.catalog_item.name}: {line.quantity}")
        
        activities_str = "; ".join(activities) if activities else "-"
        
        # Write row
        row_data = [
            timesheet.date.strftime("%d.%m.%Y") if timesheet.date else "-",
            timesheet.user.full_name if timesheet.user else "-",
            timesheet.user.employee_code if timesheet.user else "-",
            timesheet.user.role if timesheet.user else "-",
            timesheet.site.name if timesheet.site else "-",
            timesheet.check_in.strftime("%H:%M") if timesheet.check_in else "-",
            timesheet.check_out.strftime("%H:%M") if timesheet.check_out else "-",
            timesheet.break_duration or 0,
            round(hours_worked, 2),
            activities_str
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Align numbers to right
            if col_num in [8, 9]:  # Break duration, Total hours
                cell.alignment = Alignment(horizontal="right")
        
        row_num += 1
    
    # Total row
    if timesheets:
        total_row = row_num
        ws.cell(row=total_row, column=1).value = "TOTAL"
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=9).value = round(total_hours, 2)
        ws.cell(row=total_row, column=9).font = Font(bold=True)
        
        for col_num in range(1, 11):
            ws.cell(row=total_row, column=col_num).border = border
            ws.cell(row=total_row, column=col_num).fill = PatternFill(
                start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
            )
    
    # Auto-adjust column widths
    column_widths = {
        1: 12,  # Data
        2: 20,  # Angajat
        3: 15,  # Cod
        4: 15,  # Rol
        5: 25,  # Șantier
        6: 10,  # Intrare
        7: 10,  # Ieșire
        8: 12,  # Pauză
        9: 12,  # Ore Total
        10: 40  # Activități
    }
    
    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename
    filename = f"pontaje_{date_from or 'all'}_{date_to or 'all'}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/timesheets/preview")
async def preview_timesheets(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    employee_id: Optional[str] = Query(None),
    site_id: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin)
):
    """
    Preview timesheet data before export
    """
    # Build query
    query = db.query(Timesheet).join(User).join(ConstructionSite)
    
    # Apply filters
    filters = []
    if date_from:
        filters.append(Timesheet.date >= datetime.strptime(date_from, "%Y-%m-%d").date())
    if date_to:
        filters.append(Timesheet.date <= datetime.strptime(date_to, "%Y-%m-%d").date())
    if employee_id:
        filters.append(Timesheet.user_id == employee_id)
    if site_id:
        filters.append(Timesheet.site_id == site_id)
    
    if filters:
        query = query.filter(and_(*filters))
    
    # Get total count
    total = query.count()
    
    # Order and paginate
    query = query.order_by(Timesheet.date.desc(), User.full_name)
    timesheets = query.offset((page - 1) * page_size).limit(page_size).all()
    
    # Format response
    results = []
    total_hours = 0.0
    
    for timesheet in timesheets:
        # Calculate hours
        hours_worked = 0.0
        if timesheet.check_in and timesheet.check_out:
            delta = timesheet.check_out - timesheet.check_in
            hours_worked = delta.total_seconds() / 3600
            if timesheet.break_duration:
                hours_worked -= timesheet.break_duration / 60
        
        total_hours += hours_worked
        
        results.append({
            "id": timesheet.id,
            "date": timesheet.date.isoformat() if timesheet.date else None,
            "employee_name": timesheet.user.full_name if timesheet.user else None,
            "employee_code": timesheet.user.employee_code if timesheet.user else None,
            "role": timesheet.user.role if timesheet.user else None,
            "site_name": timesheet.site.name if timesheet.site else None,
            "check_in": timesheet.check_in.strftime("%H:%M") if timesheet.check_in else None,
            "check_out": timesheet.check_out.strftime("%H:%M") if timesheet.check_out else None,
            "break_duration": timesheet.break_duration or 0,
            "hours_worked": round(hours_worked, 2)
        })
    
    return {
        "timesheets": results,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_hours": round(total_hours, 2)
    }
