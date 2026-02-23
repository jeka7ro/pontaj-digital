"""
Admin API endpoints for user management
Includes: CRUD, ID card upload with OCR (easyocr), Excel import/export, avatar extraction
"""
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from typing import List, Optional
from pydantic import BaseModel, Field
from datetime import datetime, date
import hashlib
import os
import uuid
import re
import io

from app.database import get_db
from app.models import User, Role, Admin
from app.api.admin_auth import get_current_admin
from app.storage import upload_file, delete_file, get_content_type

router = APIRouter(prefix="/admin/users", tags=["admin-users"])

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads", "id_cards")
AVATAR_DIR = os.path.join(BASE_DIR, "uploads", "avatars")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(AVATAR_DIR, exist_ok=True)


# =================== PYDANTIC SCHEMAS ===================

class UserCreate(BaseModel):
    employee_code: str = Field(..., min_length=3, max_length=20)
    last_name: str = Field(..., min_length=1, max_length=100)
    first_name: str = Field(..., min_length=1, max_length=100)
    role_id: str
    pin: str = Field(..., min_length=4, max_length=6)
    birth_date: Optional[str] = None
    cnp: Optional[str] = Field(None, min_length=13, max_length=13)
    birth_place: Optional[str] = None
    id_card_series: Optional[str] = None
    phone: Optional[str] = Field(None, max_length=20)
    email: Optional[str] = None
    address: Optional[str] = None
    is_active: bool = True


class UserUpdate(BaseModel):
    last_name: Optional[str] = Field(None, min_length=1, max_length=100)
    first_name: Optional[str] = Field(None, min_length=1, max_length=100)
    full_name: Optional[str] = Field(None, min_length=2, max_length=200)
    role_id: Optional[str] = None
    is_active: Optional[bool] = None
    birth_date: Optional[str] = None
    cnp: Optional[str] = Field(None, min_length=13, max_length=13)
    birth_place: Optional[str] = None
    id_card_series: Optional[str] = None
    phone: Optional[str] = Field(None, max_length=20)
    email: Optional[str] = None
    address: Optional[str] = None


class UserPinReset(BaseModel):
    new_pin: str = Field(..., min_length=4, max_length=6)


class UserResponse(BaseModel):
    id: str
    employee_code: str
    full_name: str
    last_name: Optional[str] = None
    first_name: Optional[str] = None
    role_id: str
    role_name: str
    is_active: bool
    created_at: datetime
    last_login: Optional[datetime] = None
    birth_date: Optional[str] = None
    cnp: Optional[str] = None
    birth_place: Optional[str] = None
    id_card_series: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    avatar_path: Optional[str] = None
    id_card_path: Optional[str] = None
    contract_path: Optional[str] = None

    class Config:
        from_attributes = True


class UsersListResponse(BaseModel):
    users: List[UserResponse]
    total: int
    page: int
    page_size: int


# =================== HELPER FUNCTIONS ===================

def hash_pin(pin: str) -> str:
    return hashlib.sha256(pin.encode()).hexdigest()


def split_full_name(full_name: str):
    """Split full_name into last_name and first_name"""
    parts = full_name.strip().split(' ', 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    return full_name, ''


def build_user_response(user, role_name=None):
    """Build UserResponse from a User model instance"""
    rname = role_name or (user.role.name if user.role else 'N/A')
    last_name, first_name = split_full_name(user.full_name)
    return UserResponse(
        id=user.id,
        employee_code=user.employee_code,
        full_name=user.full_name,
        last_name=last_name,
        first_name=first_name,
        role_id=user.role_id,
        role_name=rname,
        is_active=user.is_active,
        created_at=user.created_at,
        last_login=getattr(user, 'last_login', None),
        birth_date=str(user.birth_date) if user.birth_date else None,
        cnp=user.cnp,
        birth_place=getattr(user, 'birth_place', None),
        id_card_series=getattr(user, 'id_card_series', None),
        phone=user.phone,
        email=user.email,
        address=user.address,
        avatar_path=user.avatar_path,
        id_card_path=user.id_card_path,
        contract_path=getattr(user, 'contract_path', None)
    )


def extract_id_card_data(image_path: str) -> dict:
    """
    Extract data from Romanian ID card (Carte de Identitate) using EasyOCR.
    Extracts: Nume, Prenume, CNP, Data nașterii, Loc naștere, Serie+Număr, Domiciliu.
    Also extracts avatar photo from the ID card.
    """
    result = {
        "last_name": None,
        "first_name": None,
        "cnp": None,
        "birth_date": None,
        "birth_place": None,
        "id_card_series": None,
        "address": None,
        "avatar_path": None,
        "raw_text": None,
        "success": False,
        "message": ""
    }

    try:
        from PIL import Image
        img = Image.open(image_path)
        
        # ===== Extract avatar (face photo) FIRST — works without easyocr =====
        try:
            w, h = img.size
            # Romanian ID card: photo is on the left side
            left = int(w * 0.02)
            top = int(h * 0.15)
            right = int(w * 0.35)
            bottom = int(h * 0.85)
            
            face_crop = img.crop((left, top, right, bottom))
            
            avatar_filename = f"avatar_{uuid.uuid4().hex[:8]}.jpg"
            avatar_buf = io.BytesIO()
            face_crop.save(avatar_buf, "JPEG", quality=90)
            avatar_bytes = avatar_buf.getvalue()
            avatar_url = upload_file(avatar_bytes, f"avatars/{avatar_filename}", "image/jpeg")
            result["avatar_path"] = avatar_url
        except Exception as e:
            print(f"Avatar extraction failed: {e}")
        
        # ===== Try OCR if easyocr is available =====
        try:
            import easyocr
            import numpy as np

            reader = easyocr.Reader(['ro', 'en'], gpu=False)
            img_np = np.array(img)

            # Run OCR
            ocr_results = reader.readtext(img_np, detail=1)
            
            # Extract all text
            texts = []
            for (bbox, text, prob) in ocr_results:
                texts.append(text.strip())
            
            full_text = '\n'.join(texts)
            result["raw_text"] = full_text
            
            # ===== Extract CNP (13 digits starting with 1,2,5,6,8) =====
            cnp_match = re.search(r'\b([12568]\d{12})\b', full_text.replace(' ', ''))
            if not cnp_match:
                for t in texts:
                    cleaned = t.replace(' ', '').replace('O', '0').replace('o', '0')
                    m = re.search(r'([12568]\d{12})', cleaned)
                    if m:
                        cnp_match = m
                        break
            
            if cnp_match:
                result["cnp"] = cnp_match.group(1)
                cnp = result["cnp"]
                century = '19' if cnp[0] in '12' else '20' if cnp[0] in '56' else '19'
                year = century + cnp[1:3]
                month = cnp[3:5]
                day = cnp[5:7]
                try:
                    result["birth_date"] = f"{year}-{month}-{day}"
                except Exception:
                    pass
            
            # ===== Extract Serie + Număr (pattern: XX 123456) =====
            for t in texts:
                series_match = re.search(r'\b([A-Z]{2})\s*(\d{6})\b', t)
                if series_match:
                    result["id_card_series"] = f"{series_match.group(1)} {series_match.group(2)}"
                    break
            
            # ===== Extract names from MRZ line =====
            mrz_surname = None
            mrz_firstname = None
            for t in texts:
                cleaned = t.replace(' ', '').upper()
                mrz_match = re.search(r'IDROU[A-Z]*?([A-Z]{2,})<<([A-Z]+)', cleaned)
                if not mrz_match:
                    mrz_match = re.search(r'IDROU([A-ZĂÂÎȘȚ]{2,})<<([A-ZĂÂÎȘȚ]+)', cleaned)
                if mrz_match:
                    mrz_surname = mrz_match.group(1).replace('<', '').strip()
                    mrz_firstname = mrz_match.group(2).replace('<', '').strip()
                    break
            
            if not mrz_surname:
                for t in texts:
                    cleaned = t.replace(' ', '').upper()
                    if cleaned.startswith('IDROU') and '<<' in cleaned:
                        after_idrou = cleaned[5:]
                        parts = after_idrou.split('<<')
                        name_parts = [p.replace('<', '').strip() for p in parts if p.replace('<', '').strip()]
                        if len(name_parts) >= 2:
                            mrz_surname = name_parts[0]
                            mrz_firstname = name_parts[1]
                            break
                        elif len(name_parts) == 1 and len(name_parts[0]) > 3:
                            mrz_surname = name_parts[0]
                            break

            if mrz_surname:
                result["last_name"] = mrz_surname.title()
            if mrz_firstname:
                result["first_name"] = mrz_firstname.title()
            
            # ===== Fallback: Extract fields based on label detection =====
            for i, text in enumerate(texts):
                text_upper = text.upper().strip()
                
                if not result["last_name"] and ('NUME' in text_upper or 'SURNAME' in text_upper or text_upper == 'NUME/SURNAME') and 'PRENUME' not in text_upper:
                    for j in range(i + 1, min(i + 3, len(texts))):
                        candidate = texts[j].strip()
                        candidate_upper = candidate.upper()
                        if candidate_upper and not any(kw in candidate_upper for kw in ['PRENUME', 'FIRST', 'NAME', 'GIVEN', '/', 'LOC', 'DOMICILIU', 'CNP', 'SEX']):
                            name_val = re.sub(r'[^A-ZĂÂÎȘȚa-zăâîșț\s-]', '', candidate).strip()
                            if name_val and len(name_val) > 1:
                                result["last_name"] = name_val.title()
                                break
                
                if not result["first_name"] and ('PRENUME' in text_upper or 'FIRST NAME' in text_upper or 'GIVEN' in text_upper):
                    for j in range(i + 1, min(i + 3, len(texts))):
                        candidate = texts[j].strip()
                        candidate_upper = candidate.upper()
                        if candidate_upper and not any(kw in candidate_upper for kw in ['NUME', 'LOC', 'DOMICILIU', 'CNP', 'SEX', 'NATIONAL', 'CETĂȚENI']):
                            name_val = re.sub(r'[^A-ZĂÂÎȘȚa-zăâîșț\s-]', '', candidate).strip()
                            if name_val and len(name_val) > 1:
                                result["first_name"] = name_val.title()
                                break
                
                if 'LOC' in text_upper and ('NAȘTERE' in text_upper or 'NASTERE' in text_upper or 'BIRTH' in text_upper):
                    for j in range(i + 1, min(i + 3, len(texts))):
                        candidate = texts[j].strip()
                        candidate_upper = candidate.upper()
                        if candidate_upper and not any(kw in candidate_upper for kw in ['DOMICILIU', 'CNP', 'VALID', 'SERIE']):
                            if len(candidate) > 2:
                                result["birth_place"] = candidate.title()
                                break
                
                if 'DOMICILIU' in text_upper or 'ADDRESS' in text_upper or 'DOMICILIUL' in text_upper:
                    addr_parts = []
                    for j in range(i + 1, min(i + 5, len(texts))):
                        candidate = texts[j].strip()
                        candidate_upper = candidate.upper()
                        if any(kw in candidate_upper for kw in ['CNP', 'VALID', 'SERIE', 'IDROU', 'EMIS', 'CHIP']):
                            break
                        if len(candidate) > 2:
                            addr_parts.append(candidate)
                    if addr_parts:
                        result["address"] = ', '.join(addr_parts)
            
            result["success"] = bool(result["cnp"] or result["last_name"] or result["first_name"])
            result["message"] = "Date extrase cu succes din cartea de identitate" if result["success"] else "Nu s-au putut extrage date din imagine"
            
        except ImportError:
            # OCR not available — avatar was still extracted above
            result["success"] = bool(result["avatar_path"])
            result["message"] = "Poza de profil a fost extrasă. Completează datele manual (OCR indisponibil pe server)."
        
    except Exception as e:
        result["message"] = f"Eroare procesare: {str(e)}"

    return result


# =================== API ENDPOINTS ===================

@router.get("/", response_model=UsersListResponse)
def get_users(
    page: int = 1,
    page_size: int = 20,
    search: Optional[str] = None,
    role_id: Optional[str] = None,
    is_active: Optional[bool] = None,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Get paginated list of users with optional filters"""
    query = db.query(User).join(Role)
    if search:
        query = query.filter(or_(
            User.employee_code.ilike(f"%{search}%"),
            User.full_name.ilike(f"%{search}%")
        ))
    if role_id:
        query = query.filter(User.role_id == role_id)
    if is_active is not None:
        query = query.filter(User.is_active == is_active)
    
    total = query.count()
    offset = (page - 1) * page_size
    users = query.offset(offset).limit(page_size).all()
    return UsersListResponse(users=[build_user_response(u) for u in users], total=total, page=page, page_size=page_size)


@router.get("/stats/summary")
def get_users_stats(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Get user statistics"""
    total_users = db.query(func.count(User.id)).scalar()
    active_users = db.query(func.count(User.id)).filter(User.is_active == True).scalar()
    inactive_users = total_users - active_users
    users_by_role = db.query(Role.name, func.count(User.id).label('count')).join(User).group_by(Role.name).all()
    return {
        "total_users": total_users,
        "active_users": active_users,
        "inactive_users": inactive_users,
        "users_by_role": [{"role": role, "count": count} for role, count in users_by_role]
    }


@router.get("/next-code")
def get_next_employee_code(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Get next available employee code (EMP001, EMP002, etc.)"""
    # Find all existing EMP codes and get the highest number
    existing_codes = db.query(User.employee_code).filter(
        User.employee_code.ilike("EMP%")
    ).all()
    
    max_num = 0
    for (code,) in existing_codes:
        # Extract number from EMP### format
        match = re.search(r'EMP(\d+)', code, re.IGNORECASE)
        if match:
            num = int(match.group(1))
            if num > max_num:
                max_num = num
    
    next_num = max_num + 1
    next_code = f"EMP{next_num:03d}"
    
    return {"next_code": next_code}


@router.get("/export/excel")
def export_users_excel(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Export all users to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    users = db.query(User).join(Role).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Angajați"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ["Cod Angajat", "Nume", "Prenume", "Rol", "CNP", "Serie Buletin",
               "Data Nașterii", "Loc Naștere", "Telefon", "Email", "Adresă", "Status"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, user in enumerate(users, 2):
        last_name, first_name = split_full_name(user.full_name)
        row_data = [
            user.employee_code, last_name, first_name,
            user.role.name if user.role else '',
            user.cnp or '', getattr(user, 'id_card_series', '') or '',
            str(user.birth_date) if user.birth_date else '',
            getattr(user, 'birth_place', '') or '',
            user.phone or '', user.email or '', user.address or '',
            'Activ' if user.is_active else 'Inactiv'
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 3, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"angajati_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.post("/import/excel")
async def import_users_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Import users from Excel file"""
    from openpyxl import load_workbook

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Fișierul trebuie să fie .xlsx sau .xls")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1]] if ws.max_row > 0 else []
    header_map = {}
    for i, h in enumerate(headers):
        if h:
            h_lower = h.lower().strip()
            if 'cod' in h_lower or 'employee' in h_lower:
                header_map['employee_code'] = i
            elif h_lower in ('nume', 'last name', 'last_name', 'surname'):
                header_map['last_name'] = i
            elif h_lower in ('prenume', 'first name', 'first_name', 'given name'):
                header_map['first_name'] = i
            elif 'rol' in h_lower or 'role' in h_lower:
                header_map['role'] = i
            elif 'cnp' in h_lower:
                header_map['cnp'] = i
            elif 'serie' in h_lower or 'buletin' in h_lower:
                header_map['id_card_series'] = i
            elif 'nașterii' in h_lower or 'nasterii' in h_lower or 'birth' in h_lower and 'loc' not in h_lower:
                header_map['birth_date'] = i
            elif 'loc' in h_lower and ('naștere' in h_lower or 'nastere' in h_lower or 'birth' in h_lower):
                header_map['birth_place'] = i
            elif 'telefon' in h_lower or 'phone' in h_lower:
                header_map['phone'] = i
            elif 'email' in h_lower:
                header_map['email'] = i
            elif 'adres' in h_lower or 'address' in h_lower:
                header_map['address'] = i

    if 'employee_code' not in header_map:
        raise HTTPException(status_code=400, detail="Coloana 'Cod Angajat' nu a fost găsită")

    roles = {r.name.lower(): r for r in db.query(Role).all()}
    default_role = db.query(Role).first()
    created, updated, errors = 0, 0, []

    def get_val(row, key):
        if key not in header_map:
            return None
        idx = header_map[key]
        if idx < len(row) and row[idx]:
            return str(row[idx]).strip()
        return None

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        try:
            emp_code = get_val(row, 'employee_code')
            if not emp_code:
                continue

            last_name = get_val(row, 'last_name') or ''
            first_name = get_val(row, 'first_name') or ''
            full_name = f"{last_name} {first_name}".strip() or emp_code

            role_name = (get_val(row, 'role') or '').lower()
            role = roles.get(role_name, default_role)
            if not role:
                errors.append(f"Rândul {row_idx}: Rol invalid")
                continue

            existing = db.query(User).filter(User.employee_code == emp_code).first()
            if existing:
                existing.full_name = full_name
                for field in ['cnp', 'phone', 'email', 'address', 'id_card_series', 'birth_place']:
                    val = get_val(row, field)
                    if val:
                        setattr(existing, field, val)
                bd = get_val(row, 'birth_date')
                if bd:
                    existing.birth_date = bd
                updated += 1
            else:
                new_user = User(
                    organization_id=role.organization_id,
                    employee_code=emp_code, full_name=full_name,
                    role_id=role.id, pin_hash=hash_pin('1234'), is_active=True,
                    cnp=get_val(row, 'cnp'), phone=get_val(row, 'phone'),
                    email=get_val(row, 'email'), address=get_val(row, 'address'),
                    id_card_series=get_val(row, 'id_card_series'),
                    birth_place=get_val(row, 'birth_place'),
                )
                bd = get_val(row, 'birth_date')
                if bd:
                    new_user.birth_date = bd
                db.add(new_user)
                created += 1
        except Exception as e:
            errors.append(f"Rândul {row_idx}: {str(e)}")

    db.commit()
    return {"message": f"Import finalizat: {created} creați, {updated} actualizați", "created": created, "updated": updated, "errors": errors[:10]}


@router.get("/{user_id}", response_model=UserResponse)
def get_user(user_id: str, db: Session = Depends(get_db), current_admin: Admin = Depends(get_current_admin)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return build_user_response(user)


@router.post("/", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
def create_user(user_data: UserCreate, db: Session = Depends(get_db), current_admin: Admin = Depends(get_current_admin)):
    existing = db.query(User).filter(User.employee_code == user_data.employee_code).first()
    if existing:
        raise HTTPException(status_code=400, detail="Codul de angajat există deja")
    role = db.query(Role).filter(Role.id == user_data.role_id).first()
    if not role:
        raise HTTPException(status_code=400, detail="Rol invalid")

    full_name = f"{user_data.last_name} {user_data.first_name}".strip()
    
    # Convert birth_date string to date object for SQLite
    birth_date_val = None
    if user_data.birth_date:
        from datetime import datetime as dt
        try:
            birth_date_val = dt.strptime(user_data.birth_date, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            birth_date_val = None
    
    new_user = User(
        organization_id=role.organization_id,
        employee_code=user_data.employee_code,
        full_name=full_name, role_id=user_data.role_id,
        pin_hash=hash_pin(user_data.pin), is_active=user_data.is_active,
        birth_date=birth_date_val, cnp=user_data.cnp,
        birth_place=user_data.birth_place, id_card_series=user_data.id_card_series,
        phone=user_data.phone, email=user_data.email, address=user_data.address
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return build_user_response(new_user, role.name)


@router.put("/{user_id}", response_model=UserResponse)
def update_user(user_id: str, user_data: UserUpdate, db: Session = Depends(get_db), current_admin: Admin = Depends(get_current_admin)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    if user_data.last_name is not None or user_data.first_name is not None:
        current_last, current_first = split_full_name(user.full_name)
        last = user_data.last_name if user_data.last_name is not None else current_last
        first = user_data.first_name if user_data.first_name is not None else current_first
        user.full_name = f"{last} {first}".strip()
    elif user_data.full_name is not None:
        user.full_name = user_data.full_name

    if user_data.role_id is not None:
        role = db.query(Role).filter(Role.id == user_data.role_id).first()
        if not role:
            raise HTTPException(status_code=400, detail="Invalid role ID")
        user.role_id = user_data.role_id

    for field in ['is_active', 'cnp', 'birth_place', 'id_card_series', 'phone', 'email', 'address']:
        val = getattr(user_data, field, None)
        if val is not None:
            setattr(user, field, val)

    # Handle birth_date separately - convert string to date object for SQLite
    if user_data.birth_date is not None:
        if user_data.birth_date:
            from datetime import datetime as dt
            try:
                user.birth_date = dt.strptime(user_data.birth_date, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                user.birth_date = None
        else:
            user.birth_date = None

    db.commit()
    db.refresh(user)
    return build_user_response(user)


@router.post("/{user_id}/reset-pin")
def reset_user_pin(user_id: str, pin_data: UserPinReset, db: Session = Depends(get_db), current_admin: Admin = Depends(get_current_admin)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.pin_hash = hash_pin(pin_data.new_pin)
    db.commit()
    return {"message": "PIN resetat cu succes"}


@router.post("/{user_id}/upload-id-card")
async def upload_id_card(user_id: str, file: UploadFile = File(...), db: Session = Depends(get_db), current_admin: Admin = Depends(get_current_admin)):
    """Upload ID card image, run OCR extraction, and extract avatar photo"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    allowed = ('.jpg', '.jpeg', '.png', '.webp', '.bmp')
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed:
        raise HTTPException(status_code=400, detail=f"Format neacceptat. Acceptăm: {', '.join(allowed)}")

    filename = f"{user_id}_id_card{ext}"
    storage_path = f"id_cards/{filename}"

    content = await file.read()
    
    # Upload to storage
    id_card_url = upload_file(content, storage_path, get_content_type(filename))
    user.id_card_path = id_card_url

    # Save temp file for OCR processing
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp.write(content)
        temp_path = tmp.name

    # Run OCR on temp file
    ocr_result = extract_id_card_data(temp_path)
    
    # Clean up temp file
    try:
        os.remove(temp_path)
    except:
        pass

    # Save avatar from OCR if extracted
    if ocr_result.get("avatar_path"):
        user.avatar_path = ocr_result["avatar_path"]

    # Auto-fill empty fields from OCR
    if ocr_result.get("success"):
        if not user.cnp and ocr_result.get("cnp"):
            user.cnp = ocr_result["cnp"]
        if not user.birth_date and ocr_result.get("birth_date"):
            from datetime import datetime as dt
            try:
                user.birth_date = dt.strptime(ocr_result["birth_date"], "%Y-%m-%d").date()
            except (ValueError, TypeError):
                pass
        if not user.address and ocr_result.get("address"):
            user.address = ocr_result["address"]
        if not getattr(user, 'birth_place', None) and ocr_result.get("birth_place"):
            user.birth_place = ocr_result["birth_place"]
        if not getattr(user, 'id_card_series', None) and ocr_result.get("id_card_series"):
            user.id_card_series = ocr_result["id_card_series"]

    db.commit()

    return {
        "message": "Carte de identitate încărcată cu succes",
        "id_card_path": user.id_card_path,
        "avatar_path": user.avatar_path,
        "ocr": ocr_result
    }


@router.post("/{user_id}/upload-contract")
async def upload_contract(user_id: str, file: UploadFile = File(...), db: Session = Depends(get_db), current_admin: Admin = Depends(get_current_admin)):
    """Upload employment contract (PDF/JPG) for a user"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    allowed = ('.jpg', '.jpeg', '.png', '.pdf')
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed:
        raise HTTPException(status_code=400, detail=f"Format neacceptat. Acceptăm: {', '.join(allowed)}")

    filename = f"{user_id}_contract{ext}"
    storage_path = f"contracts/{filename}"

    content = await file.read()
    content_type = "application/pdf" if ext == ".pdf" else get_content_type(filename)
    contract_url = upload_file(content, storage_path, content_type)
    user.contract_path = contract_url
    db.commit()

    return {
        "message": "Contract încărcat cu succes",
        "contract_path": user.contract_path
    }


@router.delete("/{user_id}/contract")
async def delete_contract(user_id: str, db: Session = Depends(get_db), current_admin: Admin = Depends(get_current_admin)):
    """Delete employment contract"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.contract_path = None
    db.commit()
    return {"message": "Contract șters cu succes"}


@router.post("/ocr/extract")
async def ocr_extract_only(file: UploadFile = File(...), current_admin: Admin = Depends(get_current_admin)):
    """Extract data from ID card image without saving to a user — used for pre-filling forms"""
    allowed = ('.jpg', '.jpeg', '.png', '.webp', '.bmp')
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed:
        raise HTTPException(status_code=400, detail=f"Format neacceptat")

    temp_filename = f"temp_{uuid.uuid4().hex}{ext}"
    temp_path = os.path.join(UPLOAD_DIR, temp_filename)

    content = await file.read()
    with open(temp_path, 'wb') as f:
        f.write(content)

    try:
        ocr_result = extract_id_card_data(temp_path)
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)

    return ocr_result


@router.delete("/{user_id}")
def delete_user(user_id: str, db: Session = Depends(get_db), current_admin: Admin = Depends(get_current_admin)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.is_active = False
    db.commit()
    return {"message": "Utilizator dezactivat cu succes"}
