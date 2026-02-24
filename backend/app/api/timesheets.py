from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, timedelta, date
from app.timezone import now_ro, today_ro
from typing import List, Optional
from pydantic import BaseModel

from app.database import get_db
from app.models import (
    Timesheet, TimesheetSegment, TimesheetLine, 
    User, Site, Activity, Team, Admin, ConstructionSite,
    GeofencePause
)
from app.api.admin_auth import get_current_admin, oauth2_scheme
from app.api.auth import get_current_user

router = APIRouter()

# ============================================================================
# Pydantic Models
# ============================================================================

class ActivityInput(BaseModel):
    activity_id: str
    quantity: float
    notes: Optional[str] = None

class TimesheetCreate(BaseModel):
    date: str  # YYYY-MM-DD
    site_id: str
    check_in: str  # HH:MM
    check_out: str  # HH:MM
    break_duration: Optional[int] = 0  # minutes
    notes: Optional[str] = None
    activities: List[ActivityInput] = []

class TimesheetUpdate(BaseModel):
    check_in: Optional[str] = None
    check_out: Optional[str] = None
    break_duration: Optional[int] = None
    notes: Optional[str] = None

class ApprovalAction(BaseModel):
    reason: Optional[str] = None

# ============================================================================
# Helper Functions
# ============================================================================

def get_approver_for_user(db: Session, user: User):
    """
    Determine who should approve this user's timesheet based on role hierarchy:
    - Worker → Team Lead
    - Team Lead → Site Manager  
    - Site Manager → Auto-approved (None)
    """
    if user.role.code == "SITE_MANAGER":
        return None  # Auto-approved
    elif user.role.code == "TEAM_LEAD":
        # Find site manager
        return db.query(User).filter(
            User.organization_id == user.organization_id,
            User.role.has(code="SITE_MANAGER"),
            User.is_active == True
        ).first()
    else:  # WORKER
        # Find team lead (if user is in a team)
        # For now, find any team lead
        return db.query(User).filter(
            User.organization_id == user.organization_id,
            User.role.has(code="TEAM_LEAD"),
            User.is_active == True
        ).first()

def can_approve_timesheet(approver: User, timesheet: Timesheet, db: Session):
    """Check if approver can approve this timesheet"""
    owner = timesheet.owner_user
    
    # Site Manager can approve Team Lead timesheets
    if approver.role.code == "SITE_MANAGER" and owner.role.code == "TEAM_LEAD":
        return True
    
    # Team Lead can approve Worker timesheets
    if approver.role.code == "TEAM_LEAD" and owner.role.code == "WORKER":
        return True
    
    return False

# ============================================================================
# Employee Endpoints
# ============================================================================

@router.get("/timesheets/")
async def list_my_timesheets(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List my timesheets"""
    query = db.query(Timesheet).filter(
        Timesheet.owner_user_id == current_user.id
    )
    
    # Apply filters
    if date_from:
        query = query.filter(Timesheet.date >= datetime.strptime(date_from, "%Y-%m-%d").date())
    if date_to:
        query = query.filter(Timesheet.date <= datetime.strptime(date_to, "%Y-%m-%d").date())
    if status:
        query = query.filter(Timesheet.status == status)
    
    # Get total count
    total = query.count()
    
    # Order and paginate
    timesheets = query.order_by(Timesheet.date.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    # Format response
    results = []
    for ts in timesheets:
        # Get segments and calculate hours
        segments = db.query(TimesheetSegment).filter(TimesheetSegment.timesheet_id == ts.id).all()
        total_hours = 0.0
        for seg in segments:
            if seg.check_in_time and seg.check_out_time:
                delta = seg.check_out_time - seg.check_in_time
                hours = delta.total_seconds() / 3600
                if seg.break_start_time and seg.break_end_time:
                    break_delta = seg.break_end_time - seg.break_start_time
                    hours -= break_delta.total_seconds() / 3600
                # Subtract geofence pause time
                geo_pauses = db.query(GeofencePause).filter(GeofencePause.segment_id == seg.id).all()
                for gp in geo_pauses:
                    gp_end = gp.pause_end or now_ro()
                    hours -= (gp_end - gp.pause_start).total_seconds() / 3600
                total_hours += max(0, hours)
        
        # Get activities count
        activities_count = db.query(TimesheetLine).filter(
            TimesheetLine.timesheet_id == ts.id
        ).count()
        
        results.append({
            "id": ts.id,
            "date": ts.date.isoformat() if ts.date else None,
            "status": ts.status,
            "total_hours": round(total_hours, 2),
            "activities_count": activities_count,
            "note_text": ts.note_text,
            "created_at": ts.created_at.isoformat() if ts.created_at else None
        })
    
    return {
        "timesheets": results,
        "total": total,
        "page": page,
        "page_size": page_size
    }


@router.get("/timesheets/{timesheet_id}")
async def get_timesheet_details(
    timesheet_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get timesheet details with activities"""
    timesheet = db.query(Timesheet).filter(Timesheet.id == timesheet_id).first()
    
    if not timesheet:
        raise HTTPException(status_code=404, detail="Timesheet not found")
    
    # Check ownership or approval rights
    if timesheet.owner_user_id != current_user.id:
        if not can_approve_timesheet(current_user, timesheet, db):
            raise HTTPException(status_code=403, detail="Not authorized to view this timesheet")
    
    # Get segments
    segments = db.query(TimesheetSegment).filter(
        TimesheetSegment.timesheet_id == timesheet_id
    ).all()
    
    segments_data = []
    for seg in segments:
        # Get activities for this segment
        lines = db.query(TimesheetLine).filter(
            TimesheetLine.segment_id == seg.id
        ).all()
        
        activities = []
        for line in lines:
            activity = db.query(Activity).filter(Activity.id == line.activity_id).first()
            activities.append({
                "id": line.id,
                "activity_id": line.activity_id,
                "activity_name": activity.name if activity else None,
                "quantity": float(line.quantity_numeric),
                "unit_type": line.unit_type
            })
        
        site = db.query(Site).filter(Site.id == seg.site_id).first()
        
        segments_data.append({
            "id": seg.id,
            "site_id": seg.site_id,
            "site_name": site.name if site else None,
            "check_in": seg.check_in_time.isoformat() if seg.check_in_time else None,
            "check_out": seg.check_out_time.isoformat() if seg.check_out_time else None,
            "break_start": seg.break_start_time.isoformat() if seg.break_start_time else None,
            "break_end": seg.break_end_time.isoformat() if seg.break_end_time else None,
            "activities": activities
        })
    
    return {
        "id": timesheet.id,
        "date": timesheet.date.isoformat() if timesheet.date else None,
        "status": timesheet.status,
        "note_text": timesheet.note_text,
        "segments": segments_data,
        "created_at": timesheet.created_at.isoformat() if timesheet.created_at else None,
        "updated_at": timesheet.updated_at.isoformat() if timesheet.updated_at else None
    }


@router.post("/timesheets/")
async def create_timesheet(
    data: TimesheetCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create new timesheet"""
    # Parse date
    timesheet_date = datetime.strptime(data.date, "%Y-%m-%d").date()
    
    # Check for duplicate (one timesheet per day per user)
    existing = db.query(Timesheet).filter(
        Timesheet.owner_user_id == current_user.id,
        Timesheet.date == timesheet_date
    ).first()
    
    if existing:
        raise HTTPException(status_code=400, detail="Timesheet already exists for this date")
    
    # Create timesheet
    timesheet = Timesheet(
        organization_id=current_user.organization_id,
        date=timesheet_date,
        owner_type="USER",
        owner_user_id=current_user.id,
        team_category="NO_TEAM",
        status="DRAFT",
        note_text=data.notes
    )
    db.add(timesheet)
    db.flush()
    
    # Create segment
    check_in_dt = datetime.strptime(f"{data.date} {data.check_in}", "%Y-%m-%d %H:%M")
    check_out_dt = datetime.strptime(f"{data.date} {data.check_out}", "%Y-%m-%d %H:%M")
    
    segment = TimesheetSegment(
        timesheet_id=timesheet.id,
        site_id=data.site_id,
        check_in_time=check_in_dt,
        check_out_time=check_out_dt
    )
    
    # Add break if specified
    if data.break_duration and data.break_duration > 0:
        # Assume break starts at 12:00
        break_start = datetime.strptime(f"{data.date} 12:00", "%Y-%m-%d %H:%M")
        break_end = break_start + timedelta(minutes=data.break_duration)
        segment.break_start_time = break_start
        segment.break_end_time = break_end
    
    db.add(segment)
    db.flush()
    
    # Add activities
    for activity_input in data.activities:
        activity = db.query(Activity).filter(Activity.id == activity_input.activity_id).first()
        if not activity:
            continue
        
        line = TimesheetLine(
            timesheet_id=timesheet.id,
            segment_id=segment.id,
            activity_id=activity_input.activity_id,
            quantity_numeric=activity_input.quantity,
            unit_type=activity.unit_type
        )
        db.add(line)
    
    db.commit()
    db.refresh(timesheet)
    
    return {
        "id": timesheet.id,
        "status": timesheet.status,
        "date": timesheet.date.isoformat(),
        "message": "Timesheet created successfully"
    }


@router.put("/timesheets/{timesheet_id}")
async def update_timesheet(
    timesheet_id: str,
    data: TimesheetUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update timesheet (only if DRAFT)"""
    timesheet = db.query(Timesheet).filter(
        Timesheet.id == timesheet_id,
        Timesheet.owner_user_id == current_user.id
    ).first()
    
    if not timesheet:
        raise HTTPException(status_code=404, detail="Timesheet not found")
    
    if timesheet.status != "DRAFT":
        raise HTTPException(status_code=400, detail="Can only edit DRAFT timesheets")
    
    # Update segment times
    segment = db.query(TimesheetSegment).filter(
        TimesheetSegment.timesheet_id == timesheet_id
    ).first()
    
    if segment:
        if data.check_in:
            segment.check_in_time = datetime.strptime(
                f"{timesheet.date.isoformat()} {data.check_in}", "%Y-%m-%d %H:%M"
            )
        if data.check_out:
            segment.check_out_time = datetime.strptime(
                f"{timesheet.date.isoformat()} {data.check_out}", "%Y-%m-%d %H:%M"
            )
        if data.break_duration is not None:
            if data.break_duration > 0:
                break_start = datetime.strptime(f"{timesheet.date.isoformat()} 12:00", "%Y-%m-%d %H:%M")
                break_end = break_start + timedelta(minutes=data.break_duration)
                segment.break_start_time = break_start
                segment.break_end_time = break_end
            else:
                segment.break_start_time = None
                segment.break_end_time = None
    
    if data.notes is not None:
        timesheet.note_text = data.notes
    
    db.commit()
    
    return {"message": "Timesheet updated successfully"}


@router.post("/timesheets/{timesheet_id}/submit")
async def submit_timesheet(
    timesheet_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Submit timesheet for approval"""
    timesheet = db.query(Timesheet).filter(
        Timesheet.id == timesheet_id,
        Timesheet.owner_user_id == current_user.id
    ).first()
    
    if not timesheet:
        raise HTTPException(status_code=404, detail="Timesheet not found")
    
    if timesheet.status != "DRAFT":
        raise HTTPException(status_code=400, detail="Can only submit DRAFT timesheets")
    
    # Check if has activities
    activities_count = db.query(TimesheetLine).filter(
        TimesheetLine.timesheet_id == timesheet_id
    ).count()
    
    if activities_count == 0:
        raise HTTPException(status_code=400, detail="Cannot submit timesheet without activities")
    
    # Check if Site Manager (auto-approve)
    if current_user.role.code == "SITE_MANAGER":
        timesheet.status = "APPROVED"
        timesheet.locked_at = datetime.utcnow()
        timesheet.locked_by_user_id = current_user.id
    else:
        timesheet.status = "SUBMITTED"
    
    db.commit()
    
    return {
        "status": timesheet.status,
        "message": "Timesheet submitted successfully" if timesheet.status == "SUBMITTED" else "Timesheet auto-approved"
    }


@router.delete("/timesheets/{timesheet_id}")
async def delete_timesheet(
    timesheet_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete timesheet (only if DRAFT)"""
    timesheet = db.query(Timesheet).filter(
        Timesheet.id == timesheet_id,
        Timesheet.owner_user_id == current_user.id
    ).first()
    
    if not timesheet:
        raise HTTPException(status_code=404, detail="Timesheet not found")
    
    if timesheet.status != "DRAFT":
        raise HTTPException(status_code=400, detail="Can only delete DRAFT timesheets")
    
    db.delete(timesheet)
    db.commit()
    
    return {"message": "Timesheet deleted successfully"}


# ============================================================================
# Manager/Admin Endpoints
# ============================================================================

@router.get("/admin/timesheets/pending")
async def list_pending_timesheets(
    status: Optional[str] = Query("SUBMITTED"),
    site_id: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """List pending timesheets for approval (admin sees all)"""
    query = db.query(Timesheet)
    if status:
        query = query.filter(Timesheet.status == status)
    
    # Apply filters
    if date_from:
        query = query.filter(Timesheet.date >= datetime.strptime(date_from, "%Y-%m-%d").date())
    if date_to:
        query = query.filter(Timesheet.date <= datetime.strptime(date_to, "%Y-%m-%d").date())
    
    total = query.count()
    
    timesheets = query.order_by(Timesheet.date.desc()).offset((page - 1) * page_size).limit(page_size).all()
    
    results = []
    for ts in timesheets:
        owner = db.query(User).filter(User.id == ts.owner_user_id).first()
        
        # Get total hours
        segments = db.query(TimesheetSegment).filter(TimesheetSegment.timesheet_id == ts.id).all()
        total_hours = 0.0
        for seg in segments:
            if seg.check_in_time and seg.check_out_time:
                delta = seg.check_out_time - seg.check_in_time
                hours = delta.total_seconds() / 3600
                if seg.break_start_time and seg.break_end_time:
                    break_delta = seg.break_end_time - seg.break_start_time
                    hours -= break_delta.total_seconds() / 3600
                # Subtract geofence pause time
                geo_pauses = db.query(GeofencePause).filter(GeofencePause.segment_id == seg.id).all()
                for gp in geo_pauses:
                    gp_end = gp.pause_end or now_ro()
                    hours -= (gp_end - gp.pause_start).total_seconds() / 3600
                total_hours += max(0, hours)
        
        # Get activities
        activities_count = db.query(TimesheetLine).filter(
            TimesheetLine.timesheet_id == ts.id
        ).count()
        
        results.append({
            "id": ts.id,
            "date": ts.date.isoformat() if ts.date else None,
            "employee_name": owner.full_name if owner else None,
            "employee_code": owner.employee_code if owner else None,
            "total_hours": round(total_hours, 2),
            "activities_count": activities_count,
            "note_text": ts.note_text
        })
    
    return {
        "timesheets": results,
        "total": total,
        "page": page,
        "page_size": page_size
    }


@router.post("/admin/timesheets/{timesheet_id}/approve")
async def approve_timesheet(
    timesheet_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    """Approve timesheet"""
    timesheet = db.query(Timesheet).filter(Timesheet.id == timesheet_id).first()
    
    if not timesheet:
        raise HTTPException(status_code=404, detail="Timesheet not found")
    
    if timesheet.status != "SUBMITTED":
        raise HTTPException(status_code=400, detail="Can only approve SUBMITTED timesheets")
    
    # Check approval rights
    if not can_approve_timesheet(current_user, timesheet, db):
        raise HTTPException(status_code=403, detail="Not authorized to approve this timesheet")
    
    timesheet.status = "APPROVED"
    timesheet.locked_at = datetime.utcnow()
    timesheet.locked_by_user_id = current_user.id
    
    db.commit()
    
    return {
        "status": "APPROVED",
        "message": "Timesheet approved successfully"
    }


@router.post("/admin/timesheets/{timesheet_id}/reject")
async def reject_timesheet(
    timesheet_id: str,
    action: ApprovalAction,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    """Reject timesheet with reason"""
    timesheet = db.query(Timesheet).filter(Timesheet.id == timesheet_id).first()
    
    if not timesheet:
        raise HTTPException(status_code=404, detail="Timesheet not found")
    
    if timesheet.status != "SUBMITTED":
        raise HTTPException(status_code=400, detail="Can only reject SUBMITTED timesheets")
    
    # Check approval rights
    if not can_approve_timesheet(current_user, timesheet, db):
        raise HTTPException(status_code=403, detail="Not authorized to reject this timesheet")
    
    timesheet.status = "DRAFT"  # Return to draft for editing
    timesheet.unlock_reason = action.reason or "Rejected by manager"
    timesheet.unlocked_at = datetime.utcnow()
    timesheet.unlocked_by_user_id = current_user.id
    
    db.commit()
    
    return {
        "status": "DRAFT",
        "message": "Timesheet rejected and returned to draft"
    }

@router.get("/admin/timesheets/stats")
async def get_timesheet_stats(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Get timesheet statistics for admin dashboard"""
    from datetime import datetime, timedelta
    
    today = today_ro()
    week_start = today - timedelta(days=today.weekday())
    now = now_ro()
    
    # Count all timesheets today
    today_count = db.query(Timesheet).filter(
        Timesheet.date == today,
        Timesheet.owner_type == "USER"
    ).count()
    
    # Total hours this week (ALL timesheets)
    week_timesheets = db.query(Timesheet).filter(
        Timesheet.date >= week_start,
        Timesheet.owner_type == "USER"
    ).all()
    
    total_hours_week = 0
    for ts in week_timesheets:
        segments = db.query(TimesheetSegment).filter(
            TimesheetSegment.timesheet_id == ts.id
        ).all()
        for seg in segments:
            end_time = seg.check_out_time or now
            hours = (end_time - seg.check_in_time).total_seconds() / 3600
            if seg.break_start_time:
                break_end = seg.break_end_time or now
                break_hours = (break_end - seg.break_start_time).total_seconds() / 3600
                hours -= break_hours
            total_hours_week += max(0, hours)
    
    total_users = db.query(User).filter(User.is_active == True).count()
    total_sites = db.query(ConstructionSite).count()
    
    return {
        "pending": today_count,
        "approved_today": today_count,
        "total_hours_week": round(total_hours_week, 1),
        "rejection_rate": 0,
        "total_users": total_users,
        "total_sites": total_sites
    }


@router.get("/admin/dashboard-stats")
async def get_dashboard_stats(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Comprehensive dashboard stats with daily breakdown for charts"""
    from datetime import datetime, timedelta
    
    today = today_ro()
    now = now_ro()
    
    # Last 7 days breakdown
    daily_data = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        day_name = day.strftime("%a")
        day_label = day.strftime("%d/%m")
        
        day_timesheets = db.query(Timesheet).filter(
            Timesheet.date == day,
            Timesheet.owner_type == "USER"
        ).all()
        
        day_hours = 0
        day_workers = len(day_timesheets)
        
        for ts in day_timesheets:
            segments = db.query(TimesheetSegment).filter(
                TimesheetSegment.timesheet_id == ts.id
            ).all()
            for seg in segments:
                end_time = seg.check_out_time or now
                hours = (end_time - seg.check_in_time).total_seconds() / 3600
                if seg.break_start_time:
                    break_end = seg.break_end_time or now
                    bh = (break_end - seg.break_start_time).total_seconds() / 3600
                    hours -= bh
                day_hours += max(0, hours)
        
        daily_data.append({
            "day": day_name,
            "date": day_label,
            "hours": round(day_hours, 1),
            "workers": day_workers
        })
    
    # Today's hourly breakdown
    hourly_data = []
    today_timesheets = db.query(Timesheet).filter(
        Timesheet.date == today,
        Timesheet.owner_type == "USER"
    ).all()
    
    all_segs = []
    for ts in today_timesheets:
        segs = db.query(TimesheetSegment).filter(
            TimesheetSegment.timesheet_id == ts.id
        ).all()
        all_segs.extend(segs)
    
    for hour in range(6, 21):
        hour_time = datetime(today.year, today.month, today.day, hour, 0)
        active = 0
        for seg in all_segs:
            end_t = seg.check_out_time or now
            if seg.check_in_time <= hour_time < end_t:
                on_break = False
                if seg.break_start_time:
                    be = seg.break_end_time or now
                    if seg.break_start_time <= hour_time < be:
                        on_break = True
                if not on_break:
                    active += 1
        hourly_data.append({"hour": f"{hour}:00", "workers": active})
    
    # Activity breakdown today
    activity_data = {}
    for ts in today_timesheets:
        lines = db.query(TimesheetLine).filter(
            TimesheetLine.timesheet_id == ts.id
        ).all()
        for line in lines:
            act = db.query(Activity).filter(Activity.id == line.activity_id).first()
            if act:
                if act.name not in activity_data:
                    activity_data[act.name] = {"quantity": 0, "unit_type": line.unit_type or "buc"}
                activity_data[act.name]["quantity"] += float(line.quantity_numeric) if line.quantity_numeric else 0
    
    activities_list = [{"name": k, "quantity": round(v["quantity"], 1), "unit_type": v["unit_type"]} for k, v in activity_data.items()]
    
    # Site breakdown
    site_data = {}
    for seg in all_segs:
        if seg.site_id:
            site = db.query(ConstructionSite).filter(ConstructionSite.id == seg.site_id).first()
            if site:
                site_data.setdefault(site.name, 0)
                site_data[site.name] += 1
    sites_list = [{"name": k, "workers": v} for k, v in site_data.items()]
    
    return {
        "daily": daily_data,
        "hourly": hourly_data,
        "activities": activities_list,
        "sites": sites_list,
        "today_workers": len(today_timesheets),
        "week_total_hours": round(sum(d["hours"] for d in daily_data), 1)
    }



# ============================================================================
# Live Monitoring
# ============================================================================

@router.get("/admin/timesheets/active-workers")
async def get_active_workers(
    target_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Get all workers who have a timesheet for a given date (defaults to today)"""
    from datetime import datetime
    
    query_date = date.fromisoformat(target_date) if target_date else today_ro()
    
    # Find all timesheets for the given date
    today_timesheets = db.query(Timesheet).filter(
        Timesheet.date == query_date,
        Timesheet.owner_type == "USER"
    ).all()
    
    active_workers = []
    
    for ts in today_timesheets:
        # Get the worker
        worker = db.query(User).filter(User.id == ts.owner_user_id).first()
        if not worker:
            continue
        
        # Get ALL segments for this timesheet, ordered by check_in
        all_segments = db.query(TimesheetSegment).filter(
            TimesheetSegment.timesheet_id == ts.id
        ).order_by(TimesheetSegment.check_in_time.asc()).all()
        
        if not all_segments:
            continue
        
        first_segment = all_segments[0]
        last_segment = all_segments[-1]
        
        # Get the site from the first segment
        site = db.query(ConstructionSite).filter(
            ConstructionSite.id == first_segment.site_id
        ).first()
        
        now = now_ro()
        is_on_break = False
        is_outside_geofence = False
        total_worked = 0
        total_break = 0
        total_geofence_pause = 0
        
        # Aggregate hours from ALL segments
        all_checked_out = True
        for seg in all_segments:
            if seg.check_out_time:
                seg_hours = (seg.check_out_time - seg.check_in_time).total_seconds() / 3600
            else:
                seg_hours = (now - seg.check_in_time).total_seconds() / 3600
                all_checked_out = False
            
            seg_break = 0
            if seg.break_start_time:
                if seg.break_end_time:
                    seg_break = (seg.break_end_time - seg.break_start_time).total_seconds() / 3600
                else:
                    seg_break = (now - seg.break_start_time).total_seconds() / 3600
                    is_on_break = True
            
            # Calculate geofence pause time for this segment
            seg_geofence = 0
            geo_pauses = db.query(GeofencePause).filter(
                GeofencePause.segment_id == seg.id
            ).all()
            for gp in geo_pauses:
                gp_end = gp.pause_end or now
                seg_geofence += (gp_end - gp.pause_start).total_seconds() / 3600
            
            # Check if currently outside geofence (active pause on this segment)
            active_geo = db.query(GeofencePause).filter(
                GeofencePause.segment_id == seg.id,
                GeofencePause.pause_end == None
            ).first()
            if active_geo and not seg.check_out_time:
                is_outside_geofence = True
            
            total_worked += max(0, seg_hours - max(0, seg_break) - max(0, seg_geofence))
            total_break += max(0, seg_break)
            total_geofence_pause += max(0, seg_geofence)
        
        # Detect GPS loss on active segment
        gps_lost = False
        if not all_checked_out and last_segment and not last_segment.check_out_time:
            if last_segment.last_ping_at:
                since_last_ping = (now - last_segment.last_ping_at).total_seconds()
                gps_lost = since_last_ping > 120  # 2 minutes
            else:
                since_checkin = (now - last_segment.check_in_time).total_seconds()
                gps_lost = since_checkin > 120
        
        # Determine status
        if all_checked_out:
            status = "terminat"
        elif is_outside_geofence:
            status = "geofence"
        elif gps_lost:
            status = "gps_pierdut"
        elif is_on_break:
            status = "pauză"
        else:
            status = "activ"
        
        # Get activities for this timesheet
        activity_lines = db.query(TimesheetLine).filter(
            TimesheetLine.timesheet_id == ts.id
        ).all()
        
        activity_list = []
        for tl in activity_lines:
            act = db.query(Activity).filter(Activity.id == tl.activity_id).first()
            if act:
                activity_list.append({
                    "name": act.name,
                    "quantity": float(tl.quantity_numeric) if tl.quantity_numeric else 0,
                    "unit_type": tl.unit_type,
                    "added_at": str(tl.created_at) if tl.created_at else None
                })
        
        active_workers.append({
            "worker_id": worker.id,
            "worker_name": worker.full_name,
            "employee_code": worker.employee_code,
            "avatar_path": worker.avatar_path,
            "site_name": site.name if site else "Necunoscut",
            "site_id": site.id if site else None,
            "check_in_time": str(first_segment.check_in_time),
            "check_out_time": str(last_segment.check_out_time) if last_segment.check_out_time else None,
            "worked_hours": round(total_worked, 2),
            "break_hours": round(total_break, 2),
            "geofence_pause_hours": round(total_geofence_pause, 2),
            "is_on_break": is_on_break,
            "is_outside_geofence": is_outside_geofence,
            "gps_lost": gps_lost,
            "status": status,
            "latitude": first_segment.check_in_latitude,
            "longitude": first_segment.check_in_longitude,
            "activities": activity_list
        })
    
    return {
        "active_workers": active_workers,
        "total_active": len([w for w in active_workers if w["status"] != "terminat"]),
        "total_today": len(active_workers),
        "timestamp": str(now_ro())
    }


@router.get("/admin/timesheets/worker/{worker_id}/history")
async def get_worker_history(
    worker_id: str,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Get a worker's personal info and full timesheet history"""
    from app.models import Role
    
    worker = db.query(User).filter(User.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    role = db.query(Role).filter(Role.id == worker.role_id).first()
    
    # Build timesheet query
    ts_query = db.query(Timesheet).filter(
        Timesheet.owner_user_id == worker_id,
        Timesheet.owner_type == "USER"
    )
    
    if date_from:
        ts_query = ts_query.filter(Timesheet.date >= date.fromisoformat(date_from))
    if date_to:
        ts_query = ts_query.filter(Timesheet.date <= date.fromisoformat(date_to))
    
    timesheets = ts_query.order_by(Timesheet.date.desc()).limit(60).all()
    
    history = []
    total_worked_all = 0
    total_days = len(timesheets)
    
    for ts in timesheets:
        segments = db.query(TimesheetSegment).filter(
            TimesheetSegment.timesheet_id == ts.id
        ).order_by(TimesheetSegment.check_in_time.asc()).all()
        
        if not segments:
            continue
        
        first_seg = segments[0]
        last_seg = segments[-1]
        
        site = db.query(ConstructionSite).filter(
            ConstructionSite.id == first_seg.site_id
        ).first()
        
        now = now_ro()
        total_worked = 0
        total_break = 0
        is_on_break = False
        all_checked_out = True
        
        for seg in segments:
            if seg.check_out_time:
                seg_hours = (seg.check_out_time - seg.check_in_time).total_seconds() / 3600
            else:
                seg_hours = (now - seg.check_in_time).total_seconds() / 3600
                all_checked_out = False
            
            seg_break = 0
            if seg.break_start_time:
                if seg.break_end_time:
                    seg_break = (seg.break_end_time - seg.break_start_time).total_seconds() / 3600
                else:
                    seg_break = (now - seg.break_start_time).total_seconds() / 3600
                    is_on_break = True
            
            total_worked += max(0, seg_hours - max(0, seg_break))
            total_break += max(0, seg_break)
        
        if all_checked_out:
            status = "terminat"
        elif is_on_break:
            status = "pauză"
        else:
            status = "activ"
        
        # Activities
        activity_lines = db.query(TimesheetLine).filter(
            TimesheetLine.timesheet_id == ts.id
        ).all()
        
        activities = []
        for tl in activity_lines:
            act = db.query(Activity).filter(Activity.id == tl.activity_id).first()
            if act:
                activities.append({
                    "name": act.name,
                    "quantity": float(tl.quantity_numeric) if tl.quantity_numeric else 0,
                    "unit_type": tl.unit_type,
                    "added_at": str(tl.created_at) if tl.created_at else None
                })
        
        total_worked_all += total_worked
        
        history.append({
            "date": str(ts.date),
            "site_name": site.name if site else "Necunoscut",
            "check_in": str(first_seg.check_in_time) if first_seg.check_in_time else None,
            "check_out": str(last_seg.check_out_time) if last_seg.check_out_time else None,
            "worked_hours": round(total_worked, 2),
            "break_hours": round(total_break, 2),
            "status": status,
            "activities": activities
        })
    
    return {
        "worker": {
            "id": worker.id,
            "full_name": worker.full_name,
            "employee_code": worker.employee_code,
            "phone": worker.phone,
            "email": worker.email,
            "address": worker.address,
            "birth_date": str(worker.birth_date) if worker.birth_date else None,
            "cnp": worker.cnp,
            "avatar_path": worker.avatar_path,
            "role_name": role.name if role else "Muncitor",
            "is_active": worker.is_active,
            "created_at": str(worker.created_at)
        },
        "history": history,
        "summary": {
            "total_days": total_days,
            "total_hours": round(total_worked_all, 2)
        }
    }


# ============================================================================
# Activity Management
# ============================================================================

@router.get("/activities/")
async def list_activities(
    is_active: bool = Query(True),
    db: Session = Depends(get_db),
    token: str = Depends(oauth2_scheme)
):
    """List all active activities - supports both worker and admin auth"""
    from jose import jwt, JWTError
    from app.api.admin_auth import SECRET_KEY as ADMIN_SECRET_KEY, ALGORITHM as ADMIN_ALGORITHM
    from app.config import settings
    
    organization_id = None
    user_id = None
    
    # Try worker JWT first
    try:
        payload = jwt.decode(token, settings.JWT_SECRET_KEY, algorithms=[settings.JWT_ALGORITHM])
        user_id = payload.get("sub")
    except JWTError:
        # Try admin JWT
        try:
            payload = jwt.decode(token, ADMIN_SECRET_KEY, algorithms=[ADMIN_ALGORITHM])
            user_id = payload.get("sub")
        except JWTError:
            raise HTTPException(status_code=401, detail="Invalid token")
    
    if user_id:
        # Try as worker
        user = db.query(User).filter(User.id == user_id).first()
        if user:
            organization_id = user.organization_id
        else:
            # Try as admin
            admin = db.query(Admin).filter(Admin.id == user_id).first()
            if admin:
                organization_id = admin.organization_id
    
    if not organization_id:
        raise HTTPException(status_code=401, detail="Could not determine organization")
    
    # Get categories
    from app.models import ActivityCategory
    categories = db.query(ActivityCategory).filter(
        ActivityCategory.organization_id == organization_id,
        ActivityCategory.is_active == True
    ).order_by(ActivityCategory.sort_order).all()
    
    activities = db.query(Activity).filter(
        Activity.organization_id == organization_id,
        Activity.is_active == is_active
    ).order_by(Activity.sort_order).all()
    
    # Build grouped response
    grouped = []
    for cat in categories:
        cat_activities = [a for a in activities if a.category_id == cat.id]
        if not cat_activities:
            continue
        grouped.append({
            "id": cat.id,
            "name": cat.name,
            "color": cat.color,
            "sort_order": cat.sort_order,
            "activities": [
                {
                    "id": act.id,
                    "name": act.name,
                    "description": act.description,
                    "unit_type": act.unit_type,
                    "quantity_rules": act.quantity_rules,
                    "sort_order": act.sort_order,
                    "is_active": act.is_active,
                    "category_id": cat.id,
                    "category_name": cat.name,
                    "category_color": cat.color,
                }
                for act in cat_activities
            ]
        })
    
    # Also include uncategorized activities
    uncategorized = [a for a in activities if not a.category_id]
    if uncategorized:
        grouped.append({
            "id": None,
            "name": "Alte activități",
            "color": "#94a3b8",
            "sort_order": 999,
            "activities": [
                {
                    "id": act.id,
                    "name": act.name,
                    "description": act.description,
                    "unit_type": act.unit_type,
                    "quantity_rules": act.quantity_rules,
                    "sort_order": act.sort_order or 0,
                    "is_active": act.is_active,
                    "category_id": None,
                    "category_name": "Alte activități",
                    "category_color": "#94a3b8",
                }
                for act in uncategorized
            ]
        })
    
    # Flat list for backward compatibility
    flat_activities = [
        {
            "id": act.id,
            "name": act.name,
            "description": act.description,
            "unit_type": act.unit_type,
            "quantity_rules": act.quantity_rules,
            "is_active": act.is_active,
            "category_id": act.category_id,
            "category_name": act.category.name if act.category else None,
            "category_color": act.category.color if act.category else None,
        }
        for act in activities
    ]
    
    return {
        "categories": grouped,
        "activities": flat_activities
    }


@router.post("/timesheets/{timesheet_id}/activities")
async def add_activity_to_timesheet(
    timesheet_id: str,
    activity: ActivityInput,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Add activity to timesheet"""
    timesheet = db.query(Timesheet).filter(
        Timesheet.id == timesheet_id,
        Timesheet.owner_user_id == current_user.id
    ).first()
    
    if not timesheet:
        raise HTTPException(status_code=404, detail="Timesheet not found")
    
    if timesheet.status != "DRAFT":
        raise HTTPException(status_code=400, detail="Can only add activities to DRAFT timesheets")
    
    # Get first segment
    segment = db.query(TimesheetSegment).filter(
        TimesheetSegment.timesheet_id == timesheet_id
    ).first()
    
    if not segment:
        raise HTTPException(status_code=400, detail="No segment found for this timesheet")
    
    # Get activity
    act = db.query(Activity).filter(Activity.id == activity.activity_id).first()
    if not act:
        raise HTTPException(status_code=404, detail="Activity not found")
    
    # Create line
    line = TimesheetLine(
        timesheet_id=timesheet_id,
        segment_id=segment.id,
        activity_id=activity.activity_id,
        quantity_numeric=activity.quantity,
        unit_type=act.unit_type
    )
    db.add(line)
    db.commit()
    db.refresh(line)
    
    return {
        "id": line.id,
        "message": "Activity added successfully"
    }


@router.delete("/timesheets/activities/{activity_id}")
async def delete_activity_from_timesheet(
    activity_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete activity from timesheet"""
    line = db.query(TimesheetLine).filter(TimesheetLine.id == activity_id).first()
    
    if not line:
        raise HTTPException(status_code=404, detail="Activity not found")
    
    # Check ownership
    timesheet = db.query(Timesheet).filter(
        Timesheet.id == line.timesheet_id,
        Timesheet.owner_user_id == current_user.id
    ).first()
    
    if not timesheet:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    if timesheet.status != "DRAFT":
        raise HTTPException(status_code=400, detail="Can only delete activities from DRAFT timesheets")
    
    db.delete(line)
    db.commit()
    
    return {"message": "Activity deleted successfully"}

# ============================================================================
# Admin Activity Management
# ============================================================================

class ActivityCreate(BaseModel):
    name: str
    unit_type: str
    category_id: Optional[str] = None
    description: Optional[str] = None
    quantity_rules: Optional[str] = None
    sort_order: Optional[int] = 0
    is_active: bool = True

class ActivityUpdate(BaseModel):
    name: Optional[str] = None
    unit_type: Optional[str] = None
    category_id: Optional[str] = None
    description: Optional[str] = None
    quantity_rules: Optional[str] = None
    sort_order: Optional[int] = None
    is_active: Optional[bool] = None

@router.post("/admin/activities/")
async def create_activity(
    data: ActivityCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    """Create new activity"""
    activity = Activity(
        organization_id=current_user.organization_id,
        name=data.name,
        unit_type=data.unit_type,
        category_id=data.category_id,
        description=data.description,
        quantity_rules=data.quantity_rules,
        sort_order=data.sort_order or 0,
        is_active=data.is_active
    )
    db.add(activity)
    db.commit()
    db.refresh(activity)
    
    return {
        "id": activity.id,
        "message": "Activity created successfully"
    }

@router.put("/admin/activities/{activity_id}")
async def update_activity(
    activity_id: str,
    data: ActivityUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    """Update activity"""
    activity = db.query(Activity).filter(
        Activity.id == activity_id,
        Activity.organization_id == current_user.organization_id
    ).first()
    
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found")
    
    if data.name is not None:
        activity.name = data.name
    if data.unit_type is not None:
        activity.unit_type = data.unit_type
    if data.category_id is not None:
        activity.category_id = data.category_id if data.category_id != "" else None
    if data.description is not None:
        activity.description = data.description
    if data.quantity_rules is not None:
        activity.quantity_rules = data.quantity_rules
    if data.sort_order is not None:
        activity.sort_order = data.sort_order
    if data.is_active is not None:
        activity.is_active = data.is_active
    
    db.commit()
    
    return {"message": "Activity updated successfully"}

@router.delete("/admin/activities/{activity_id}")
async def delete_activity(
    activity_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    """Delete (deactivate) activity"""
    activity = db.query(Activity).filter(
        Activity.id == activity_id,
        Activity.organization_id == current_user.organization_id
    ).first()
    
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found")
    
    # Check if activity is used in any timesheets
    usage_count = db.query(TimesheetLine).filter(
        TimesheetLine.activity_id == activity_id
    ).count()
    
    if usage_count > 0:
        # Don't delete, just deactivate
        activity.is_active = False
        db.commit()
        return {"message": f"Activity deactivated (used in {usage_count} timesheets)"}
    else:
        # Safe to delete
        db.delete(activity)
        db.commit()
        return {"message": "Activity deleted successfully"}

# ============================================================================
# Admin Activity Category Management
# ============================================================================

class CategoryCreate(BaseModel):
    name: str
    color: str = "#3b82f6"
    sort_order: int = 0

class CategoryUpdate(BaseModel):
    name: Optional[str] = None
    color: Optional[str] = None
    sort_order: Optional[int] = None
    is_active: Optional[bool] = None

@router.get("/admin/activity-categories/")
async def list_activity_categories(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """List all activity categories with activity counts"""
    from app.models import ActivityCategory
    
    categories = db.query(ActivityCategory).filter(
        ActivityCategory.organization_id == current_admin.organization_id
    ).order_by(ActivityCategory.sort_order).all()
    
    result = []
    for cat in categories:
        activity_count = db.query(Activity).filter(
            Activity.category_id == cat.id
        ).count()
        active_count = db.query(Activity).filter(
            Activity.category_id == cat.id,
            Activity.is_active == True
        ).count()
        result.append({
            "id": cat.id,
            "name": cat.name,
            "color": cat.color,
            "sort_order": cat.sort_order,
            "is_active": cat.is_active,
            "activity_count": activity_count,
            "active_activity_count": active_count
        })
    
    return {"categories": result}

@router.post("/admin/activity-categories/")
async def create_activity_category(
    data: CategoryCreate,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Create new activity category"""
    from app.models import ActivityCategory
    
    category = ActivityCategory(
        organization_id=current_admin.organization_id,
        name=data.name,
        color=data.color,
        sort_order=data.sort_order
    )
    db.add(category)
    db.commit()
    db.refresh(category)
    
    return {
        "id": category.id,
        "message": "Category created successfully"
    }

@router.put("/admin/activity-categories/{category_id}")
async def update_activity_category(
    category_id: str,
    data: CategoryUpdate,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Update activity category"""
    from app.models import ActivityCategory
    
    category = db.query(ActivityCategory).filter(
        ActivityCategory.id == category_id,
        ActivityCategory.organization_id == current_admin.organization_id
    ).first()
    
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    if data.name is not None:
        category.name = data.name
    if data.color is not None:
        category.color = data.color
    if data.sort_order is not None:
        category.sort_order = data.sort_order
    if data.is_active is not None:
        category.is_active = data.is_active
    
    db.commit()
    
    return {"message": "Category updated successfully"}

@router.delete("/admin/activity-categories/{category_id}")
async def delete_activity_category(
    category_id: str,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Delete activity category (moves activities to uncategorized)"""
    from app.models import ActivityCategory
    
    category = db.query(ActivityCategory).filter(
        ActivityCategory.id == category_id,
        ActivityCategory.organization_id == current_admin.organization_id
    ).first()
    
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    # Move activities to uncategorized
    activities = db.query(Activity).filter(
        Activity.category_id == category_id
    ).all()
    for act in activities:
        act.category_id = None
    
    db.delete(category)
    db.commit()
    
    return {"message": f"Category deleted. {len(activities)} activities moved to uncategorized."}


# ============================================================================
# Overtime Approval
# ============================================================================

@router.put("/admin/timesheets/segments/{segment_id}/approve-overtime")
async def approve_segment_overtime(
    segment_id: str,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Approve overtime for a timesheet segment"""
    segment = db.query(TimesheetSegment).filter(
        TimesheetSegment.id == segment_id
    ).first()
    
    if not segment:
        raise HTTPException(status_code=404, detail="Segment not found")
    
    # Verify it belongs to admin's org via timesheet
    timesheet = db.query(Timesheet).filter(
        Timesheet.id == segment.timesheet_id,
        Timesheet.organization_id == current_admin.organization_id
    ).first()
    
    if not timesheet:
        raise HTTPException(status_code=404, detail="Timesheet not found")
    
    segment.overtime_approved = True
    segment.overtime_approved_by = current_admin.id
    segment.overtime_approved_at = datetime.utcnow()
    
    db.commit()
    
    return {
        "message": f"Overtime aprobat ({segment.overtime_minutes} minute)",
        "overtime_minutes": segment.overtime_minutes,
        "approved_by": current_admin.full_name
    }


@router.get("/admin/notifications/feed")
async def get_notification_feed(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Get real-time event feed from today's activity"""
    from datetime import datetime
    
    today = today_ro()
    now = now_ro()
    
    # Fetch all today timesheets with segments
    today_timesheets = db.query(Timesheet).filter(
        Timesheet.date == today,
        Timesheet.owner_type == "USER"
    ).all()
    
    events = []
    for ts in today_timesheets:
        worker = db.query(User).filter(User.id == ts.owner_user_id).first()
        if not worker:
            continue
        wname = worker.full_name or "Necunoscut"
        avatar = worker.avatar_path
        
        segments = db.query(TimesheetSegment).filter(
            TimesheetSegment.timesheet_id == ts.id
        ).all()
        
        first_seg = segments[0] if segments else None
        site = db.query(ConstructionSite).filter(ConstructionSite.id == first_seg.site_id).first() if first_seg else None
        site_name = site.name if site else "Necunoscut"
        
        for seg in segments:
            # Check-in event
            events.append({
                "type": "check_in",
                "icon": "🟢",
                "message": f"{wname} a intrat pe șantier",
                "detail": site_name,
                "time": str(seg.check_in_time),
                "timestamp": seg.check_in_time.timestamp(),
                "worker_name": wname,
                "avatar_path": avatar
            })
            
            # Break start
            if seg.break_start_time:
                events.append({
                    "type": "break_start",
                    "icon": "☕",
                    "message": f"{wname} a intrat în pauză",
                    "detail": site_name,
                    "time": str(seg.break_start_time),
                    "timestamp": seg.break_start_time.timestamp(),
                    "worker_name": wname,
                    "avatar_path": avatar
                })
            
            # Break end
            if seg.break_end_time:
                events.append({
                    "type": "break_end",
                    "icon": "🔄",
                    "message": f"{wname} a revenit din pauză",
                    "detail": site_name,
                    "time": str(seg.break_end_time),
                    "timestamp": seg.break_end_time.timestamp(),
                    "worker_name": wname,
                    "avatar_path": avatar
                })
            
            # Check-out
            if seg.check_out_time:
                hours = (seg.check_out_time - seg.check_in_time).total_seconds() / 3600
                events.append({
                    "type": "check_out",
                    "icon": "🔴",
                    "message": f"{wname} a ieșit de pe șantier",
                    "detail": f"{site_name} — {round(hours, 1)}h",
                    "time": str(seg.check_out_time),
                    "timestamp": seg.check_out_time.timestamp(),
                    "worker_name": wname,
                    "avatar_path": avatar
                })
    
    # Sort newest first
    events.sort(key=lambda e: e["timestamp"], reverse=True)
    
    return {
        "events": events[:50],
        "total": len(events)
    }


# ============================================================================
# Excel Export — Timesheets
# ============================================================================

@router.get("/admin/timesheets/export/excel")
async def export_timesheets_excel(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    site_id: Optional[str] = None,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Export timesheets to Excel (date range, optional site filter)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    import io

    today = today_ro()
    start = date.fromisoformat(date_from) if date_from else today - timedelta(days=30)
    end = date.fromisoformat(date_to) if date_to else today

    ts_query = db.query(Timesheet).filter(
        Timesheet.date >= start,
        Timesheet.date <= end,
        Timesheet.owner_type == "USER"
    )

    timesheets = ts_query.order_by(Timesheet.date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Pontaje"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["Data", "Angajat", "Cod", "Șantier", "Check-in", "Check-out", "Ore Lucrate", "Pauză (h)", "Status", "Activități"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row = 2
    now = now_ro()
    for ts in timesheets:
        worker = db.query(User).filter(User.id == ts.owner_user_id).first()
        segments = db.query(TimesheetSegment).filter(
            TimesheetSegment.timesheet_id == ts.id
        ).order_by(TimesheetSegment.check_in_time.asc()).all()

        if not segments:
            continue

        if site_id and not any(s.site_id == site_id for s in segments):
            continue

        first_seg = segments[0]
        last_seg = segments[-1]
        site = db.query(ConstructionSite).filter(ConstructionSite.id == first_seg.site_id).first()

        total_worked = 0
        total_break = 0
        all_checked_out = True
        for seg in segments:
            if seg.check_out_time:
                seg_hours = (seg.check_out_time - seg.check_in_time).total_seconds() / 3600
            else:
                seg_hours = (now - seg.check_in_time).total_seconds() / 3600
                all_checked_out = False
            seg_break = 0
            if seg.break_start_time:
                be = seg.break_end_time or now
                seg_break = (be - seg.break_start_time).total_seconds() / 3600
            total_worked += max(0, seg_hours - max(0, seg_break))
            total_break += max(0, seg_break)

        # Activities
        lines = db.query(TimesheetLine).filter(TimesheetLine.timesheet_id == ts.id).all()
        act_strs = []
        for ln in lines:
            act = db.query(Activity).filter(Activity.id == ln.activity_id).first()
            if act:
                act_strs.append(f"{act.name}: {float(ln.quantity_numeric)} {ln.unit_type}")

        status = "Terminat" if all_checked_out else "Activ"

        ws.cell(row=row, column=1, value=str(ts.date)).border = thin_border
        ws.cell(row=row, column=2, value=worker.full_name if worker else "N/A").border = thin_border
        ws.cell(row=row, column=3, value=worker.employee_code if worker else "N/A").border = thin_border
        ws.cell(row=row, column=4, value=site.name if site else "N/A").border = thin_border
        ws.cell(row=row, column=5, value=str(first_seg.check_in_time.strftime("%H:%M")) if first_seg.check_in_time else "").border = thin_border
        ws.cell(row=row, column=6, value=str(last_seg.check_out_time.strftime("%H:%M")) if last_seg.check_out_time else "—").border = thin_border
        ws.cell(row=row, column=7, value=round(total_worked, 2)).border = thin_border
        ws.cell(row=row, column=8, value=round(total_break, 2)).border = thin_border
        ws.cell(row=row, column=9, value=status).border = thin_border
        ws.cell(row=row, column=10, value="; ".join(act_strs) if act_strs else "—").border = thin_border
        row += 1

    # Auto-width
    for col in range(1, len(headers) + 1):
        max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, row))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"pontaje_{start}_{end}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================================================
# Excel Export — Activities
# ============================================================================

@router.get("/admin/activities/export/excel")
async def export_activities_excel(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Export all activities catalog to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    from app.models import ActivityCategory
    import io

    categories = db.query(ActivityCategory).filter(
        ActivityCategory.organization_id == current_admin.organization_id
    ).order_by(ActivityCategory.sort_order).all()

    activities = db.query(Activity).filter(
        Activity.organization_id == current_admin.organization_id
    ).order_by(Activity.sort_order).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Activități"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["Categorie", "Activitate", "Descriere", "Unitate Măsură", "Ordine", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    cat_map = {c.id: c.name for c in categories}
    row = 2
    for act in activities:
        ws.cell(row=row, column=1, value=cat_map.get(act.category_id, "Necategorizat")).border = thin_border
        ws.cell(row=row, column=2, value=act.name).border = thin_border
        ws.cell(row=row, column=3, value=act.description or "").border = thin_border
        ws.cell(row=row, column=4, value=act.unit_type).border = thin_border
        ws.cell(row=row, column=5, value=act.sort_order or 0).border = thin_border
        ws.cell(row=row, column=6, value="Activ" if act.is_active else "Inactiv").border = thin_border
        row += 1

    for col in range(1, len(headers) + 1):
        max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, row))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=activitati.xlsx"}
    )
